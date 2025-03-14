from openpyxl import load_workbook
from database.crud import save_values
from database.db import Jury

async def process_excel_file(file_stream):
    workbook = load_workbook(file_stream)

    jury_sheet = workbook["Жюри"]
    for row in jury_sheet.iter_rows(min_row=2, values_only=True):
        jury_id = row[0]
        full_name = row[1]

        existing_jury = Jury.objects(jury_id=jury_id).first()
        if not existing_jury:
            jury = Jury(jury_id=jury_id, full_name=full_name)
            jury.save()

    students_sheet = workbook["Студенты"]
    potok_column_index = 4

    for row in students_sheet.iter_rows(min_row=2, values_only=True):
        student_id = row[0]
        student_name = row[1]
        topic = row[2]
        potok = f"Поток {row[potok_column_index - 1]}"
        date = row[4]
        juries_id = str(row[5]).split(" ")

        await save_values(student_id, student_name, topic, potok, date, juries_id)

    workbook.close()